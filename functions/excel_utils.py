
import os
import csv
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers

def parse_photo_date_time(name):
    try:
        base = os.path.basename(name)
        date_fmt = "--/--/----"
        heure_fmt = "--:--"

        date_match = re.search(r"(20\d{6})", base)
        time_match = re.search(r"_(\d{6})_", base)

        if date_match:
            date_raw = date_match.group(1)
            y, m, d = date_raw[:4], date_raw[4:6], date_raw[6:8]
            date_fmt = f"{d}/{m}/{y}"

        if time_match:
            time_raw = time_match.group(1)
            h, mi, s = time_raw[:2], time_raw[2:4], time_raw[4:6]
            heure_fmt = f"{h}:{mi}:{s}"

        return date_fmt, heure_fmt
    except Exception:
        return "--/--/----", "--:--"

def load_stations_info():
    station_info = {}
    with open('Database/station.csv', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            station_info[row['Station']] = {
                'Commune': row['Commune'],
                'Latitude': row['Latitude'],
                'Longitude': row['Longitude'],
                'Z_CC49': row['Z_CC49']
            }
    return station_info

def create_or_update_excel(input_folder, output_folder):
    excel_file = os.path.join(output_folder, "resultats_photos.xlsx")
    station_info = load_stations_info()

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    all_photos = {}
    for station in os.listdir(input_folder):
        station_path = os.path.join(input_folder, station)
        if not os.path.isdir(station_path):
            continue
        all_photos[station] = sorted([
            f for f in os.listdir(station_path)
            if f.lower().endswith(('.jpg', '.jpeg', '.png'))
        ])

    for station, photos in all_photos.items():
        old_data = {}
        if station in wb.sheetnames:
            ws_old = wb[station]
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                # Cas : ancienne version avec colonne "Date / Heure"
                if len(row) == 3:
                    photo, date_heure, result = row
                    if isinstance(date_heure, str) and " " in date_heure:
                        date_part, heure_part = date_heure.split(" ")
                        heure_part = heure_part.replace("h", ":").replace("m", ":")
                    else:
                        date_part, heure_part = "--/--/----", "--:--"
                    old_data[photo] = (date_part, heure_part, result)
                elif len(row) == 4:
                    photo, date_part, heure_part, result = row
                    old_data[photo] = (date_part, heure_part, result)
            wb.remove(ws_old)

        ws = wb.create_sheet(title=station)
        ws.append(["Nom de la photo", "Date", "Heure", "Résultat"])

        for photo in photos:
            date_fmt, heure_fmt = parse_photo_date_time(photo)
            if photo in old_data:
                _, _, old_result = old_data[photo]
                ws.append([photo, date_fmt, heure_fmt, old_result])
            else:
                ws.append([photo, date_fmt, heure_fmt, ""])

            last_row = ws.max_row
            ws.cell(row=last_row, column=2).number_format = numbers.FORMAT_TEXT
            ws.cell(row=last_row, column=3).number_format = numbers.FORMAT_TEXT

    resume_sheet = wb["Résumé"] if "Résumé" in wb.sheetnames else wb.create_sheet(title="Résumé")
    resume_sheet.delete_rows(1, resume_sheet.max_row)
    resume_sheet.append(["Station", "Commune", "Latitude", "Longitude", "Z_CC49", "Nombre de photos"])

    for station, photos in all_photos.items():
        info = station_info.get(station, {})
        commune = info.get('Commune', 'Inconnu')
        lat = info.get('Latitude', '')
        lon = info.get('Longitude', '')
        z_cc49 = info.get('Z_CC49', '')
        resume_sheet.append([station, commune, lat, lon, z_cc49, len(photos)])

    wb.save(excel_file)
    return excel_file

def update_excel_result(excel_file, sheet, row, new_value):
    wb = load_workbook(excel_file)
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    ws.cell(row=row + 2, column=4, value=new_value)
    wb.save(excel_file)
